"""
Contact loading and phone validation.
"""

import os
import re

from openpyxl import load_workbook

from config import CONFIG_FILE
from utils import fatal


def load_message():
    """Load common message from config.txt."""
    if not os.path.exists(CONFIG_FILE):
        fatal(f"config.txt not found at: {CONFIG_FILE}")
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        message = f.read().strip()
    if not message:
        fatal("config.txt is empty. Write your message in it.")
    return message


def validate_phone(phone):
    """Validate and normalize phone number. Returns (ok, normalized, error)."""
    phone = re.sub(r"[\s\-\.\(\)]", "", str(phone).strip())
    if not phone.startswith("+"):
        if re.match(r"^[6-9]\d{9}$", phone):
            phone = "+91" + phone
        elif re.match(r"^91[6-9]\d{9}$", phone):
            phone = "+" + phone
        else:
            return False, phone, "Invalid format: need country code (e.g., +919876543210)"
    digits = re.sub(r"\D", "", phone)
    if len(digits) < 10:
        return False, phone, f"Too short: {len(digits)} digits"
    if len(digits) > 15:
        return False, phone, f"Too long: {len(digits)} digits"
    return True, phone, None


def load_contacts(filepath):
    """Load contacts from Excel. Returns list of (name, phone)."""
    if not os.path.exists(filepath):
        fatal(f"Input file not found: {filepath}\n   Run: python generate_test_data.py")
    try:
        wb = load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        fatal(f"Cannot read Excel: {e}")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    name_col = next((i for i, h in enumerate(headers) if "name" in h), None)
    phone_col = next((i for i, h in enumerate(headers) if any(k in h for k in ("phone", "number", "mobile"))), None)
    if phone_col is None:
        fatal(f"No Phone/Number/Mobile column found. Headers: {headers}")

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[phone_col] is None:
            continue
        name = str(row[name_col]) if name_col is not None and row[name_col] else "Unknown"
        contacts.append((name, str(row[phone_col]).strip()))
    wb.close()
    return contacts
