"""
Excel status report generation.
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from config import SCRIPT_DIR, STATUS_COLORS, HEADER_COLOR


def _style_cell(cell, font=None, alignment=None, fill=None, border=None):
    """Apply styles to a cell."""
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell


def create_status_excel(batch_id, records, message):
    """Create output status report Excel file."""
    filename = f"status_report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    filepath = os.path.join(SCRIPT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Status Report"

    border = Border(*(Side(style="thin") for _ in range(4)))
    hdr_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center")

    headers = ["Sr No", "Name", "Phone", "Status", "Error Details", "Timestamp"]
    for col, header in enumerate(headers, 1):
        _style_cell(ws.cell(row=1, column=col, value=header),
                    font=hdr_font, fill=hdr_fill, alignment=hdr_align, border=border)

    for i, rec in enumerate(records, 1):
        row = i + 1
        _style_cell(ws.cell(row=row, column=1, value=i), font=data_font, alignment=center, border=border)
        _style_cell(ws.cell(row=row, column=2, value=rec["name"]), font=data_font, border=border)
        _style_cell(ws.cell(row=row, column=3, value=rec["phone"]), font=data_font, border=border)

        color = STATUS_COLORS.get(rec["status"], "FFFFFF")
        _style_cell(ws.cell(row=row, column=4, value=rec["status"]),
                    font=Font(name="Calibri", size=11, bold=True), alignment=center,
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                    border=border)

        _style_cell(ws.cell(row=row, column=5, value=rec["error_details"] or ""), font=data_font, border=border)
        _style_cell(ws.cell(row=row, column=6, value=rec["timestamp"]), font=data_font, border=border)

    msg_row = len(records) + 3
    _style_cell(ws.cell(row=msg_row, column=1, value="Common Message:"),
                font=Font(name="Calibri", size=11, bold=True))
    _style_cell(ws.cell(row=msg_row, column=2, value=message),
                font=Font(name="Calibri", size=11, italic=True))
    ws.merge_cells(start_row=msg_row, start_column=2, end_row=msg_row, end_column=6)

    for col, w in zip("ABCDEF", [8, 25, 20, 15, 40, 22]):
        ws.column_dimensions[col].width = w

    wb.save(filepath)
    return filepath
