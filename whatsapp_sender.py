"""
WhatsApp Bulk Message Sender (Selenium-based)
==============================================
Sends a common message to all contacts from an Excel file via WhatsApp Web.
Uses Selenium to interact with the browser like a human — finding elements,
clicking buttons, typing messages, and detecting errors.

Tracks status in SQLite database + output Excel report.

Usage:
  python whatsapp_sender.py                     # Send messages
  python whatsapp_sender.py --dry-run           # Test without sending
  python whatsapp_sender.py --input other.xlsx  # Different input file
  python whatsapp_sender.py --retries 3         # Retry failed messages

Requirements:
  pip install selenium webdriver-manager openpyxl
"""

import os
import sys
import re
import time
import random
import shutil
import argparse
import uuid
import subprocess
from datetime import datetime
from urllib.parse import quote

# ─── Dependency Checks ───────────────────────────────────────

MISSING_DEPS = []

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    MISSING_DEPS.append("openpyxl")

try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import (
        TimeoutException, NoSuchElementException,
        WebDriverException, StaleElementReferenceException,
    )
except ImportError:
    MISSING_DEPS.append("selenium")

try:
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    MISSING_DEPS.append("webdriver-manager")

if MISSING_DEPS:
    print(f"Missing: {', '.join(MISSING_DEPS)}")
    print(f"Run: pip install {' '.join(MISSING_DEPS)}")
    sys.exit(1)

from database import init_db, insert_record, update_record, get_all_records, get_summary

# ─── Constants ───────────────────────────────────────────────

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(SCRIPT_DIR, "config.txt")
DEFAULT_INPUT = os.path.join(SCRIPT_DIR, "contacts.xlsx")
CHROME_PROFILE_DIR = os.path.join(SCRIPT_DIR, "chrome_profile")
DELAY_BETWEEN_MESSAGES = 20
MAX_RETRIES = 2

# Status constants
STATUS_SENT = "Sent"
STATUS_FAILED = "Failed"
STATUS_NO_WA = "No WhatsApp"
STATUS_PENDING = "Pending"
STATUS_DRY_RUN = "Dry Run"

STATUS_ICONS = {
    STATUS_SENT: "[OK]",
    STATUS_FAILED: "[FAIL]",
    STATUS_NO_WA: "[NO WA]",
    STATUS_PENDING: "[...]",
    STATUS_DRY_RUN: "[TEST]",
}

STATUS_COLORS = {
    STATUS_SENT: "C8E6C9",
    STATUS_FAILED: "FFCDD2",
    STATUS_NO_WA: "FFE0B2",
    STATUS_PENDING: "E0E0E0",
    STATUS_DRY_RUN: "BBDEFB",
}

HEADER_COLOR = "1565C0"


# ─── Utility Helpers ────────────────────────────────────────

def log(msg, end="\n"):
    """Print with flush."""
    print(msg, end=end, flush=True)


def fatal(msg):
    """Print error and exit."""
    log(f"ERROR: {msg}")
    sys.exit(1)


def human_delay(min_s=1.0, max_s=3.0):
    """Sleep a random human-like delay."""
    time.sleep(random.uniform(min_s, max_s))


def check_internet():
    """Check internet connectivity."""
    try:
        result = subprocess.run(
            ["ping", "-n", "1", "-w", "3000", "8.8.8.8"],
            capture_output=True, timeout=5,
        )
        return result.returncode == 0
    except Exception:
        return False


# ─── Data Loading ────────────────────────────────────────────

def load_message():
    """Load common message from config.txt."""
    if not os.path.exists(CONFIG_FILE):
        fatal(f"config.txt not found at: {CONFIG_FILE}")
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        message = f.read().strip()
    if not message:
        fatal("config.txt is empty. Write your message in it.")
    return message


def validate_phone(phone):
    """Validate and normalize phone number. Returns (ok, normalized, error)."""
    phone = re.sub(r"[\s\-\.\(\)]", "", str(phone).strip())
    if not phone.startswith("+"):
        if re.match(r"^[6-9]\d{9}$", phone):
            phone = "+91" + phone
        elif re.match(r"^91[6-9]\d{9}$", phone):
            phone = "+" + phone
        else:
            return False, phone, "Invalid format: need country code (e.g., +919876543210)"
    digits = re.sub(r"\D", "", phone)
    if len(digits) < 10:
        return False, phone, f"Too short: {len(digits)} digits"
    if len(digits) > 15:
        return False, phone, f"Too long: {len(digits)} digits"
    return True, phone, None


def load_contacts(filepath):
    """Load contacts from Excel. Returns list of (name, phone)."""
    if not os.path.exists(filepath):
        fatal(f"Input file not found: {filepath}\n   Run: python generate_test_data.py")
    try:
        wb = load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        fatal(f"Cannot read Excel: {e}")

    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    name_col = next((i for i, h in enumerate(headers) if "name" in h), None)
    phone_col = next((i for i, h in enumerate(headers) if any(k in h for k in ("phone", "number", "mobile"))), None)
    if phone_col is None:
        fatal(f"No Phone/Number/Mobile column found. Headers: {headers}")

    contacts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[phone_col] is None:
            continue
        name = str(row[name_col]) if name_col is not None and row[name_col] else "Unknown"
        contacts.append((name, str(row[phone_col]).strip()))
    wb.close()
    return contacts


# ─── Selenium WhatsApp Driver ───────────────────────────────

class WhatsAppDriver:
    """Manages Chrome browser session for WhatsApp Web."""

    def __init__(self):
        self.driver = None

    def start(self):
        """Launch Chrome with WhatsApp Web profile (persists login)."""
        log("Starting Chrome browser...")
        options = Options()
        options.add_argument(f"--user-data-dir={CHROME_PROFILE_DIR}")
        options.add_argument("--profile-directory=Default")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-extensions")
        options.add_argument("--start-maximized")
        # Suppress automation detection
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)

        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(service=service, options=options)
        except WebDriverException as e:
            fatal(f"Cannot start Chrome: {e}")

        # Navigate to WhatsApp Web
        log("Opening WhatsApp Web...")
        self.driver.get("https://web.whatsapp.com")

        # Wait for WhatsApp to fully load (QR scan or auto-login)
        log("Waiting for WhatsApp Web to load (scan QR code if prompted)...")
        try:
            # Wait up to 60 seconds for the search/chat area to appear
            WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div[contenteditable="true"][data-tab="3"]'))
            )
            log("WhatsApp Web loaded successfully!")
        except TimeoutException:
            fatal("WhatsApp Web did not load in 60 seconds. Please scan QR code and try again.")

    def is_alive(self):
        """Check if the browser session is still active."""
        try:
            _ = self.driver.title
            return True
        except Exception:
            return False

    def restart_if_needed(self):
        """Restart browser if it crashed or was closed."""
        if not self.is_alive():
            log("Browser session lost. Restarting...")
            try:
                self.driver.quit()
            except Exception:
                pass
            self.start()
            return True
        return False

    def send_message(self, phone, message):
        """
        Send a message to a phone number via WhatsApp Web.
        Returns (status, error_details)
        """
        # Navigate to chat with this phone number
        url = f"https://web.whatsapp.com/send?phone={phone}&text={quote(message)}"
        self.driver.get(url)

        # Wait for the page to process - detect message box or error
        try:
            # Wait for either the message input box or an "invalid phone" popup
            WebDriverWait(self.driver, 25).until(
                lambda d: self._find_message_box(d) or self._detect_invalid_number(d)
            )
        except TimeoutException:
            return STATUS_FAILED, "Page did not load in time"

        # Check if number is invalid / not on WhatsApp
        if self._detect_invalid_number(self.driver):
            # Close any popup
            self._dismiss_popup()
            return STATUS_NO_WA, "Phone number is not on WhatsApp"

        # Find the message input box
        msg_box = self._find_message_box(self.driver)
        if not msg_box:
            return STATUS_FAILED, "Could not find message input box"

        # Human-like: small pause before sending
        human_delay(1.0, 2.0)

        # Press Enter to send (message is already pre-filled from URL)
        try:
            msg_box.send_keys(Keys.ENTER)
        except Exception as e:
            return STATUS_FAILED, f"Could not press Enter: {e}"

        # Wait for message to be sent (check for sent tick)
        human_delay(3.0, 5.0)

        return STATUS_SENT, None

    def _find_message_box(self, driver):
        """Find the WhatsApp message input box."""
        try:
            # The message input area (contenteditable div for typing)
            elements = driver.find_elements(By.CSS_SELECTOR, 'div[contenteditable="true"][data-tab="10"]')
            if elements:
                return elements[0]
            # Fallback selector
            elements = driver.find_elements(By.CSS_SELECTOR, 'div[contenteditable="true"][title="Type a message"]')
            if elements:
                return elements[0]
            # Another fallback
            footer = driver.find_elements(By.CSS_SELECTOR, "footer div[contenteditable='true']")
            if footer:
                return footer[0]
        except (NoSuchElementException, StaleElementReferenceException):
            pass
        return None

    def _detect_invalid_number(self, driver):
        """Detect if WhatsApp shows 'invalid number' or 'not on WhatsApp' popup."""
        try:
            # Look for error dialog / popup
            error_selectors = [
                "div[data-animate-modal-popup='true']",
                "div._3J6wB",  # popup container
            ]
            for selector in error_selectors:
                elements = driver.find_elements(By.CSS_SELECTOR, selector)
                for el in elements:
                    text = el.text.lower()
                    if any(k in text for k in ["invalid", "not on whatsapp", "phone number shared via url"]):
                        return True

            # Also check the page body for error text
            body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
            if "phone number shared via url is invalid" in body_text:
                return True
        except Exception:
            pass
        return False

    def _dismiss_popup(self):
        """Try to close any popup/dialog."""
        try:
            ok_buttons = self.driver.find_elements(By.CSS_SELECTOR, "div[role='button']")
            for btn in ok_buttons:
                if btn.text.strip().lower() in ("ok", "close"):
                    btn.click()
                    human_delay(0.5, 1.0)
                    return
        except Exception:
            pass

    def quit(self):
        """Close the browser."""
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass


# ─── Message Processing ─────────────────────────────────────

def process_contact(wa_driver, name, phone, message, batch_id, dry_run, retries):
    """Process a single contact. Returns (status, error)."""
    is_valid, normalized, error = validate_phone(phone)
    if not is_valid:
        rid = insert_record(name, phone, batch_id)
        update_record(rid, STATUS_FAILED, error)
        return STATUS_FAILED, error

    rid = insert_record(name, normalized, batch_id)

    if dry_run:
        update_record(rid, STATUS_DRY_RUN, None)
        return STATUS_DRY_RUN, None

    # Retry loop
    last_error = None
    for attempt in range(1, retries + 1):
        try:
            # Restart browser if crashed
            wa_driver.restart_if_needed()

            status, error = wa_driver.send_message(normalized, message)

            if status == STATUS_SENT or status == STATUS_NO_WA:
                update_record(rid, status, error)
                return status, error

            # Failed — retry if possible
            last_error = error
            if attempt < retries:
                log(f"\n   Attempt {attempt} failed: {error}. Retrying in 10s...")
                time.sleep(10)
        except KeyboardInterrupt:
            update_record(rid, STATUS_FAILED, "Interrupted by user")
            raise
        except Exception as e:
            last_error = str(e)
            if attempt < retries:
                log(f"\n   Attempt {attempt} error: {last_error}. Retrying in 10s...")
                time.sleep(10)

    update_record(rid, STATUS_FAILED, f"{last_error} (after {retries} attempts)")
    return STATUS_FAILED, last_error


# ─── Excel Report ────────────────────────────────────────────

def _style_cell(cell, font=None, alignment=None, fill=None, border=None):
    """Apply styles to a cell."""
    if font: cell.font = font
    if alignment: cell.alignment = alignment
    if fill: cell.fill = fill
    if border: cell.border = border
    return cell


def create_status_excel(batch_id, records, message):
    """Create output status report Excel file."""
    filename = f"status_report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    filepath = os.path.join(SCRIPT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Status Report"

    border = Border(*(Side(style="thin") for _ in range(4)))
    hdr_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    hdr_fill = PatternFill(start_color=HEADER_COLOR, end_color=HEADER_COLOR, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=11)
    center = Alignment(horizontal="center")

    headers = ["Sr No", "Name", "Phone", "Status", "Error Details", "Timestamp"]
    for col, header in enumerate(headers, 1):
        _style_cell(ws.cell(row=1, column=col, value=header),
                    font=hdr_font, fill=hdr_fill, alignment=hdr_align, border=border)

    for i, rec in enumerate(records, 1):
        row = i + 1
        _style_cell(ws.cell(row=row, column=1, value=i), font=data_font, alignment=center, border=border)
        _style_cell(ws.cell(row=row, column=2, value=rec["name"]), font=data_font, border=border)
        _style_cell(ws.cell(row=row, column=3, value=rec["phone"]), font=data_font, border=border)

        color = STATUS_COLORS.get(rec["status"], "FFFFFF")
        _style_cell(ws.cell(row=row, column=4, value=rec["status"]),
                    font=Font(name="Calibri", size=11, bold=True), alignment=center,
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                    border=border)

        _style_cell(ws.cell(row=row, column=5, value=rec["error_details"] or ""), font=data_font, border=border)
        _style_cell(ws.cell(row=row, column=6, value=rec["timestamp"]), font=data_font, border=border)

    msg_row = len(records) + 3
    _style_cell(ws.cell(row=msg_row, column=1, value="Common Message:"),
                font=Font(name="Calibri", size=11, bold=True))
    _style_cell(ws.cell(row=msg_row, column=2, value=message),
                font=Font(name="Calibri", size=11, italic=True))
    ws.merge_cells(start_row=msg_row, start_column=2, end_row=msg_row, end_column=6)

    for col, w in zip("ABCDEF", [8, 25, 20, 15, 40, 22]):
        ws.column_dimensions[col].width = w

    wb.save(filepath)
    return filepath


# ─── Main ────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="WhatsApp Bulk Message Sender")
    parser.add_argument("--dry-run", action="store_true", help="Test without sending")
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Input Excel file")
    parser.add_argument("--retries", type=int, default=MAX_RETRIES, help="Retry attempts")
    args = parser.parse_args()

    log("=" * 60)
    log("   WhatsApp Bulk Message Sender (Selenium)")
    log("=" * 60)
    if args.dry_run:
        log("   MODE: DRY RUN")
    log("")

    # Pre-flight
    if not args.dry_run and not check_internet():
        fatal("No internet connection.")

    # Load data
    message = load_message()
    log(f"Message: \"{message[:80]}{'...' if len(message) > 80 else ''}\"")
    contacts = load_contacts(args.input)
    log(f"Loaded {len(contacts)} contacts from: {os.path.basename(args.input)}")

    if not contacts:
        log("No contacts found.")
        sys.exit(0)

    # Init DB
    init_db()
    batch_id = str(uuid.uuid4())[:8]
    log(f"Batch: {batch_id}")
    log("-" * 60)

    # Start browser (skip in dry-run)
    wa_driver = WhatsAppDriver()
    if not args.dry_run:
        wa_driver.start()

    # Process
    interrupted = False
    for i, (name, phone) in enumerate(contacts, 1):
        log(f"\n[{i}/{len(contacts)}] {name} ({phone}) ... ", end="")

        try:
            status, error = process_contact(
                wa_driver, name, phone, message, batch_id, args.dry_run, args.retries
            )
        except KeyboardInterrupt:
            log("INTERRUPTED!")
            interrupted = True
            break

        icon = STATUS_ICONS.get(status, "[?]")
        log(f"{icon} {status}", end="")
        if error:
            log(f" - {error}", end="")
        log("")

        # Human-like delay between messages
        if not args.dry_run and i < len(contacts):
            delay = DELAY_BETWEEN_MESSAGES + random.uniform(-3, 5)
            log(f"   Waiting {delay:.0f}s before next...")
            try:
                time.sleep(delay)
            except KeyboardInterrupt:
                log("\nStopped by user.")
                interrupted = True
                break

    # Cleanup browser
    wa_driver.quit()

    # Report
    log("\n" + "=" * 60)
    records = get_all_records(batch_id)
    try:
        report = create_status_excel(batch_id, records, message)
        log(f"Report: {os.path.basename(report)}")
    except Exception as e:
        log(f"Warning: Could not create report: {e}")

    summary = get_summary(batch_id)
    log("\nSummary:")
    log(f"  Total:          {len(contacts)}")
    for status, count in summary.items():
        log(f"  {STATUS_ICONS.get(status, '[?]')} {status:14s} {count}")

    log(f"\nDB: whatsapp_sender.db (Batch: {batch_id})")
    if interrupted:
        log("Note: Interrupted. Run again to continue.")
    log("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n\nTerminated by user.")
        sys.exit(0)
    except Exception as e:
        log(f"\nFatal error: {e}")
        sys.exit(1)
