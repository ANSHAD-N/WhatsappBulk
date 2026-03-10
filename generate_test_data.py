"""
Generate test Excel file with random Indian phone numbers for testing.
Creates contacts.xlsx with Name and Phone columns.
"""

import random
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("Error: openpyxl is not installed.")
    print("Run: pip install openpyxl")
    exit(1)

# Random Indian first names for test data
FIRST_NAMES = [
    "Aarav", "Vivaan", "Aditya", "Vihaan", "Arjun",
    "Sai", "Reyansh", "Ayaan", "Krishna", "Ishaan",
    "Ananya", "Diya", "Priya", "Riya", "Saanvi",
    "Meera", "Kavya", "Neha", "Pooja", "Shreya"
]

LAST_NAMES = [
    "Sharma", "Patel", "Kumar", "Singh", "Reddy",
    "Gupta", "Verma", "Joshi", "Nair", "Iyer",
    "Desai", "Mehta", "Shah", "Rao", "Mishra"
]


def generate_random_phone():
    """Generate a random Indian phone number with +91 prefix."""
    # Indian mobile numbers start with 6, 7, 8, or 9
    first_digit = random.choice([6, 7, 8, 9])
    remaining = "".join([str(random.randint(0, 9)) for _ in range(9)])
    return f"+91{first_digit}{remaining}"


def generate_random_name():
    """Generate a random full name."""
    return f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"


def create_test_excel(filename="contacts.xlsx", num_contacts=10):
    """Create a test Excel file with random contacts."""
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # --- Styling ---
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=11)

    # --- Headers ---
    headers = ["Sr No", "Name", "Phone"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data rows ---
    for i in range(1, num_contacts + 1):
        row = i + 1
        name = generate_random_name()
        phone = generate_random_phone()

        ws.cell(row=row, column=1, value=i).font = data_font
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = thin_border

        ws.cell(row=row, column=2, value=name).font = data_font
        ws.cell(row=row, column=2).border = thin_border

        ws.cell(row=row, column=3, value=phone).font = data_font
        ws.cell(row=row, column=3).border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 20

    wb.save(filepath)
    print(f"✅ Test file created: {filepath}")
    print(f"   → {num_contacts} random contacts generated")
    return filepath


if __name__ == "__main__":
    create_test_excel()
